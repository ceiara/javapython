from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws['A1'] = '123123'
ws['A2'] = '안녕하세요'
ws.title = "처음에 자동sheet"
ws.sheet_properties.tabColor = "1072BA"

ms1 = wb.create_sheet("MySheet")
ms1['B1'] = '하이하이'
ms1['B2'] = '하하하;;;'

ms2 = wb.create_sheet("MySheet",0)
ms2['C1'] = 'test'
ms2['C2'] = 'test'

ms3 = wb["MySheet"]

wb.save('a.xlsx')
wb.close() # 자원 반환