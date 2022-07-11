from openpyxl import Workbook
import numpy as np

wb = Workbook() # 새로운 xl파일 생성
ws = wb.active  # 새로운 worksheet를 가르키는 역할

index = 1
for r in range(1,11):
    for c in range(1,11):
        ws.cell(row=r,column=c).value = np.random.randint(0,100)
        # index += 1

ws.cell(row=1,column=1).value = 1

wb.save('b.xlsx')
wb.close()